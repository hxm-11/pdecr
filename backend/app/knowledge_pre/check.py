from pathlib import Path
from openpyxl import load_workbook
import re

file = Path(
    r"C:\Users\AXU6WX\Documents\full-stack-fastapi-template-master\backend\app\rag\excel_source\T0001-PDECR - eCom anorld M6 Bolt 04M-02.xlsx"
)

wb = load_workbook(file, data_only=True)

keywords = [
    "step 4",
    "step 7",
    "approval",
    "approver",
    "signature",
    "sign",
    "implementation approval",
    "technical feasibility",
    "validation plan approval",
    "签名",
    "签字",
    "批准",
    "确认",
    "development",
    "purchasing",
    "mfe",
    "tef",
    "cos",
    "quality",
    "cpjm",
    "moex",
    "log",
    "研发",
    "开发",
    "采购",
    "工艺",
    "样品",
    "质量",
    "客户项目",
    "生产",
    "物流",
]

name_patterns = [
    re.compile(r"\b[A-Z]{2,}\s+[A-Z][a-z]+\b"),      # XIANG Liangshan
    re.compile(r"\b[A-Z][a-z]+\s+[A-Z][a-z]+\b"),    # Wang Xiaolong
    re.compile(r"[\u4e00-\u9fff]{2,4}"),             # 中文名
]

for ws in wb.worksheets:
    print("\n==============================")
    print("Sheet:", ws.title)
    print("==============================")

    for r in range(1, ws.max_row + 1):
        values = [
            str(ws.cell(r, c).value or "").strip()
            for c in range(1, ws.max_column + 1)
        ]

        row_text = " | ".join(values)
        low = row_text.lower()

        has_keyword = any(k.lower() in low for k in keywords)
        has_name = any(p.search(row_text) for p in name_patterns)

        if has_keyword or has_name:
            print(f"row {r}: {row_text}")