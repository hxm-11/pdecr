from pathlib import Path
from openpyxl import load_workbook
import re
import sys

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

RAG_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = RAG_DIR / "excel_source"
OUT_DIR = RAG_DIR / "knowledge"
STRUCTURED_SECTION_HEADING = "## Structured Signature Fields"

OUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# 基础工具函数
# ============================================================

def cell_text(v) -> str:
    return str(v or "").strip()


def clean_person_name(value: str) -> str:
    value = cell_text(value)

    if not value:
        return ""

    # 统一括号
    value = value.replace("（", "(").replace("）", ")")

    # 去掉部门括号，例如 Ku Qiong (RBCD/EIS4)
    value = re.sub(r"\([^)]*\)", "", value).strip()

    # 去掉多余空格
    value = re.sub(r"\s+", " ", value).strip()

    # 去掉开头结尾奇怪符号
    value = value.strip("/:_- ")

    # 如果是全小写英文名，转成首字母大写，方便后续识别
    # xiang liangshan -> Xiang Liangshan
    if re.fullmatch(r"[a-z]+\s+[a-z]+", value):
        value = " ".join([p.capitalize() for p in value.split()])

    # 如果是 surname 全小写 + given name 全小写，也保留为首字母大写格式
    # tao jiong -> Tao Jiong
    if re.fullmatch(r"[a-z]+\s+[a-z]+(?:\s+[a-z]+)?", value):
        value = " ".join([p.capitalize() for p in value.split()])

    return value


def is_valid_person_name(value: str) -> bool:
    value = clean_person_name(value)
    low = value.lower()

    if not value:
        return False

    invalid_exact_words = {
        "x", "-", "/", "na", "n/a", "none", "null", "空白",
        "不影响", "无", "无需", "不涉及", "没有", "否",
        "工程师", "设计", "测试", "验证", "机加工", "装配测试",
        "电机设计", "控制器设计", "本体开发", "内部质量",
        "客户项目经理", "平台项目经理", "客户项目",
        "采购", "生产", "质量", "样品", "物流", "开发", "研发", "工艺",
        "核心组", "其他", "其它",
        "签名", "签字", "负责人", "确认", "批准",
    }

    if value in invalid_exact_words or low in invalid_exact_words:
        return False

    invalid_contains_words = [
        "development", "purchasing", "quality", "mfe", "tef",
        "cos", "cpjm", "moex", "log", "other", "others",
        "signature", "approval", "implementation", "responsible",
        "acc.to project", "send to work level", "folder link",
        "microsoft teams", "business plan", "capacity",
        "doc.", "document", "release", "drawing", "bom",
        "签名", "签字", "批准", "确认", "负责人",
        "不影响", "影响", "测试", "验证", "设计", "工程师",
        "装配", "机加工", "客户项目", "平台项目",
        "部门", "候选", "矩阵",
    ]

    if any(w in low or w in value for w in invalid_contains_words):
        return False

    # XIANG Liangshan / TAO Jiong / HE Yonggang
    if re.fullmatch(r"[A-Z]{2,}\s+[A-Z][a-z]+", value):
        return True

    # Xiang Liangshan / Tao Jiong / He Yonggang / Xu Baochun
    if re.fullmatch(r"[A-Z][a-z]+\s+[A-Z][a-z]+", value):
        return True

    # Wang Xiaolong / Li Pingzheng
    if re.fullmatch(r"[A-Z][a-z]+\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?", value):
        return True

    # 兼容 Fan change / Sun tie，经 clean 后一般会变成 Fan Change / Sun Tie
    if re.fullmatch(r"[A-Z][a-z]+\s+[a-z]+", value):
        return True

    # 中文姓名
    if re.fullmatch(r"[\u4e00-\u9fff]{2,4}", value):
        return True

    return False


def split_persons(cell_value: str):
    """
    一个单元格里可能有多个人：
    Ku Qiong (RBCD/EIS4)
    Fan change (RBCD/EIS4)
    Li Pingzheng (RBCD/EIS4)
    """

    text = str(cell_value or "").strip()
    if not text:
        return []

    text = text.replace("CC:/", "\n").replace("CC:", "\n")
    text = text.replace("；", ";")

    parts = re.split(r"[\n;]+", text)

    names = []

    for part in parts:
        part = str(part or "").strip()
        if not part:
            continue

        # 先尝试从整段里抓英文名
        english_names = re.findall(
            r"\b(?:[A-Z]{2,}|[A-Z][a-z]+)\s+[A-Z][a-z]+\b",
            part,
        )

        if english_names:
            for n in english_names:
                n = clean_person_name(n)
                if is_valid_person_name(n):
                    names.append(n)
            continue

        # 再清理括号后判断
        part = clean_person_name(part)

        if is_valid_person_name(part):
            names.append(part)

    return names


def unique_keep_order(items):
    seen = set()
    out = []

    for item in items:
        item = clean_person_name(item)
        if not item:
            continue

        key = item.lower()
        if key not in seen:
            seen.add(key)
            out.append(item)

    return out


def get_cell(ws, row, col) -> str:
    """
    读取单元格文本。
    如果是合并单元格，尝试读取合并区域左上角的值。
    """

    if row < 1 or col < 1 or row > ws.max_row or col > ws.max_column:
        return ""

    value = ws.cell(row=row, column=col).value
    if value is not None:
        return cell_text(value)

    current_coord = ws.cell(row=row, column=col).coordinate

    for merged_range in ws.merged_cells.ranges:
        if current_coord in merged_range:
            top_left = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value
            return cell_text(top_left)

    return ""


def has_any_valid_person(d: dict) -> bool:
    for v in d.values():
        if isinstance(v, str) and is_valid_person_name(v):
            return True
    return False


# ============================================================
# Step 4 / Step 7 实际签字人抽取
# ============================================================

APPROVAL_RESULT_KEYS = [
    "approval_development_person",
    "approval_purchasing_person",
    "approval_mfe_person",
    "approval_cos_person",
    "approval_quality_person",
    "approval_cpjm_person",
    "approval_moex_person",
    "approval_log_person",
]


DEPT_MAPPING = [
    ("approval_development_person", ["development", "研发", "开发"]),
    ("approval_purchasing_person", ["purchasing", "采购", "pps"]),
    ("approval_mfe_person", ["mfe", "tef", "工艺"]),
    ("approval_quality_person", ["quality", "质量", "qmm"]),
    ("approval_cpjm_person", ["cpjm", "客户项目", "客户项"]),
    ("approval_cos_person", ["cos", "样品"]),
    ("approval_moex_person", ["moex", "moe", "生产"]),
    ("approval_log_person", ["log", "物流"]),
]


def match_dept_key(text: str) -> str:
    text = cell_text(text)
    low = text.lower()

    for result_key, aliases in DEPT_MAPPING:
        for alias in aliases:
            if alias.lower() in low or alias in text:
                return result_key

    return ""


def find_approval_anchor_rows(ws):
    """
    找 Step 4 / Step 7 的锚点行。
    不依赖固定行号。
    """

    anchors = []

    keywords = [
        "step 4",
        "step 7",
        "technical feasibility",
        "validation plan approval",
        "implementation approval",
        "suggested approvers",
        "技术可行性",
        "验证计划批准",
        "导入清单",
        "导入确认",
        "签字",
        "签名",
    ]

    for r in range(1, ws.max_row + 1):
        row_text = " ".join(
            get_cell(ws, r, c)
            for c in range(1, ws.max_column + 1)
        ).lower()

        if any(k.lower() in row_text for k in keywords):
            anchors.append(r)

    return anchors


def find_department_headers_near_anchor(ws, anchor_row, max_scan_rows=30):
    """
    在锚点后若干行内找部门表头。
    支持：
    - Development / Purchasing / MFE / Quality / CPjM / COS
    - MOEx / LOG / Others
    - 不同行位置
    """

    header_rows = []

    end_row = min(anchor_row + max_scan_rows, ws.max_row)

    for r in range(anchor_row, end_row + 1):
        headers = []
        used_keys = set()

        for c in range(1, ws.max_column + 1):
            text = get_cell(ws, r, c)
            key = match_dept_key(text)

            if key and key not in used_keys:
                headers.append(
                    {
                        "key": key,
                        "col": c,
                        "text": text,
                    }
                )
                used_keys.add(key)

        # 一行里至少出现 2 个部门，才认为是签字表头
        if len(headers) >= 2:
            header_rows.append((r, headers))

    return header_rows


def find_person_near_column(ws, row, start_col, end_col):
    """
    在某个部门表头对应的列范围内找人名。
    适配合并单元格、列轻微偏移。
    """

    start_col = max(1, start_col)
    end_col = min(ws.max_column, end_col)

    for c in range(start_col, end_col + 1):
        value = clean_person_name(get_cell(ws, row, c))

        if is_valid_person_name(value):
            return value

    return ""


def extract_persons_under_headers(ws, header_row, headers, result):
    """
    在表头下方 1~8 行内找实际人名。
    支持表头和人名之间有空行。
    """

    for person_row in range(header_row + 1, min(header_row + 9, ws.max_row) + 1):
        row_filled = 0

        for idx, h in enumerate(headers):
            key = h["key"]
            col = h["col"]

            if result.get(key):
                continue

            # 当前部门列的搜索范围：
            # 从当前表头列开始，到下一个表头列前一列结束。
            if idx + 1 < len(headers):
                next_col = headers[idx + 1]["col"]
                start_col = col
                end_col = max(col, next_col - 1)
            else:
                start_col = col
                end_col = min(ws.max_column, col + 3)

            person = find_person_near_column(
                ws=ws,
                row=person_row,
                start_col=start_col,
                end_col=end_col,
            )

            if person:
                result[key] = person
                row_filled += 1

        if row_filled > 0:
            print(f"[DEBUG] {ws.title} row {person_row} 抽到 {row_filled} 个实际签字人")
            print("[DEBUG] 当前 actual_approval:", result)

    return result


def extract_actual_approval_from_workbook(wb):
    """
    抽取实际 Step 4 / Step 7 签字区域。
    关键原则：
    1. 跳过 Signature Matrix。
    2. 不从 Step 6.1 Responsible 里抽 approval。
    3. 支持不同文件行号不同。
    4. 支持两段式表格。
    """

    result = {key: "" for key in APPROVAL_RESULT_KEYS}

    for ws in wb.worksheets:
        sheet_name = ws.title.strip().lower()

        # 候选人矩阵不能作为实际签字人
        if sheet_name == "signature matrix":
            continue

        anchors = find_approval_anchor_rows(ws)

        if anchors:
            print(f"[DEBUG] {ws.title} approval anchors: {anchors}")

        for anchor_row in anchors:
            header_rows = find_department_headers_near_anchor(
                ws=ws,
                anchor_row=anchor_row,
                max_scan_rows=30,
            )

            for header_row, headers in header_rows:
                print(f"[DEBUG] 找到 approval 表头: sheet={ws.title}, row={header_row}")
                print("[DEBUG] headers:", headers)

                result = extract_persons_under_headers(
                    ws=ws,
                    header_row=header_row,
                    headers=headers,
                    result=result,
                )

    # 最后再清理一遍，防止“不影响”等误入
    for key in APPROVAL_RESULT_KEYS:
        value = result.get(key, "")
        if not is_valid_person_name(value):
            result[key] = ""

    return result


# ============================================================
# Step 6.1 Implementation Responsible 抽取
# ============================================================

def extract_step61_responsible_from_workbook(wb):
    """
    抽取 Step 6.1 Implementation check list 里的 Responsible。
    注意：这不是 approval person，只能作为 implementation responsible。
    """

    result = {
        "implementation_development_responsible": "",
        "implementation_purchasing_responsible": "",
        "implementation_mfe_responsible": "",
        "implementation_cos_responsible": "",
        "implementation_quality_responsible": "",
        "implementation_cpjm_responsible": "",
        "implementation_moex_responsible": "",
        "implementation_log_responsible": "",
    }

    dept_map = {
        "development": "implementation_development_responsible",
        "purchasing": "implementation_purchasing_responsible",
        "pps": "implementation_purchasing_responsible",
        "mfe": "implementation_mfe_responsible",
        "tef": "implementation_mfe_responsible",
        "manufacturing": "implementation_mfe_responsible",
        "cos": "implementation_cos_responsible",
        "quality": "implementation_quality_responsible",
        "qmm": "implementation_quality_responsible",
        "cpjm": "implementation_cpjm_responsible",
        "pm": "implementation_cpjm_responsible",
        "moex": "implementation_moex_responsible",
        "moe": "implementation_moex_responsible",
        "log": "implementation_log_responsible",
    }

    for ws in wb.worksheets:
        if ws.title.strip().lower() != "implementation":
            continue

        for r in range(1, ws.max_row + 1):
            values = [
                get_cell(ws, r, c)
                for c in range(1, ws.max_column + 1)
            ]

            row_text = " ".join(values).lower()

            matched_key = ""

            for dept, key in dept_map.items():
                if dept in row_text:
                    matched_key = key
                    break

            if not matched_key:
                continue

            for value in values:
                names = split_persons(value)

                if names:
                    if not result[matched_key]:
                        result[matched_key] = names[0]
                    break

    return result


# ============================================================
# Signature Matrix 候选人抽取
# ============================================================

def extract_signature_matrix_candidates_from_workbook(wb):
    """
    从 Signature Matrix 中按列抽候选人。
    注意：
    - 这里只输出 candidates。
    - 不能写入 approval_xxx_person。
    """

    result = {
        "development_candidates": [],
        "purchasing_candidates": [],
        "mfe_candidates": [],
        "cos_candidates": [],
        "quality_candidates": [],
        "cpjm_candidates": [],
        "moex_candidates": [],
        "log_candidates": [],
    }

    def map_header_to_key(header_text: str) -> str:
        text = str(header_text or "").strip()
        low = text.lower()

        # 顺序很重要：先具体后泛化
        if "采购" in text or "purchasing" in low:
            return "purchasing_candidates"

        if "cos" in low:
            return "cos_candidates"

        if "质量" in text or "quality" in low or "内部质量" in text:
            return "quality_candidates"

        if "客户项目" in text or "cpjm" in low or "customer" in low:
            return "cpjm_candidates"

        if "物流" in text or "log" in low:
            return "log_candidates"

        if (
            "生产" in text
            or "moe" in low
            or "moex" in low
            or "机加工" in text
            or "装配" in text
        ):
            return "moex_candidates"

        if "mfe" in low or "tef" in low or "工艺" in text:
            return "mfe_candidates"

        if (
            "本体开发" in text
            or "core group" in low
            or "电机设计" in text
            or "控制器设计" in text
            or "测试" in text
            or "验证" in text
            or "development" in low
        ):
            return "development_candidates"

        return ""

    for ws in wb.worksheets:
        if ws.title.strip().lower() != "signature matrix":
            continue

        for header_row in range(1, ws.max_row + 1):
            row_text = " ".join(
                get_cell(ws, header_row, c)
                for c in range(1, ws.max_column + 1)
            )

            # 找 Signature Matrix 表头行
            if not (
                "本体开发" in row_text
                or "采购" in row_text
                or "客户项目经理" in row_text
                or "core group" in row_text.lower()
                or "design fmea" in row_text.lower()
            ):
                continue

            col_to_key = {}

            # 尝试组合两层表头：header_row + header_row + 1
            for c in range(1, ws.max_column + 1):
                h1 = get_cell(ws, header_row, c)
                h2 = get_cell(ws, header_row + 1, c)
                header_text = f"{h1} {h2}".strip()

                key = map_header_to_key(header_text)

                if key:
                    col_to_key[c] = key

            if not col_to_key:
                continue

            print(f"[DEBUG] Signature Matrix header_row={header_row}")
            print("[DEBUG] col_to_key:", col_to_key)

            # 一般 row34-row36 是工程师 / SM / DM
            for r in range(header_row + 2, min(header_row + 15, ws.max_row) + 1):
                for c, key in col_to_key.items():
                    value = get_cell(ws, r, c)
                    names = split_persons(value)

                    for name in names:
                        if is_valid_person_name(name):
                            result[key].append(name)

            # 找到一个矩阵后就退出当前 sheet
            break

    for key in result:
        result[key] = unique_keep_order(result[key])

    return result


# ============================================================
# Step 3.3 Affected Documents
# ============================================================

def extract_step33_affected_documents_from_workbook(wb):
    """
    Step 3.3 的 No/Yes 勾选框很多是 Excel 控件，openpyxl 读不到。
    因此这里采用保守策略：
    1. 尝试从 Implementation 的 Y/N 行反推部分字段。
    2. 读不到则默认 no，避免 LLM 后续推理成全 yes。
    """

    result = {
        "interface_fmea_value": "",
        "product_fmea_value": "",
        "special_characteristics_value": "",
        "imds_value": "",
        "offer_drawing_value": "",
        "tcd_value": "",
        "norm_wb_hf_value": "",
        "affected_document_other_value": "",
    }

    for ws in wb.worksheets:
        if ws.title.strip().lower() != "implementation":
            continue

        for r in range(1, ws.max_row + 1):
            values = [
                get_cell(ws, r, c)
                for c in range(1, ws.max_column + 1)
            ]

            row_text = " ".join(values).lower()

            yn = ""
            for v in values:
                vv = v.strip().upper()
                if vv in ["Y", "N"]:
                    yn = vv
                    break

            if not yn:
                continue

            value = "yes" if yn == "Y" else "no"

            if "d-fmea" in row_text or "dfmea" in row_text or "product fmea" in row_text:
                result["product_fmea_value"] = value

            if "offer drawing" in row_text:
                result["offer_drawing_value"] = value

            if "tcd" in row_text:
                result["tcd_value"] = value

            if "norm" in row_text or "wb" in row_text or "hf" in row_text:
                result["norm_wb_hf_value"] = value

            if "imds" in row_text:
                result["imds_value"] = value

            if "interface fmea" in row_text or "ifmea" in row_text:
                result["interface_fmea_value"] = value

            if "special characteristics" in row_text or "psc" in row_text:
                result["special_characteristics_value"] = value

            if "wi check" in row_text or "work instruction" in row_text:
                result["affected_document_other_value"] = value

    # 兜底：读不到就默认 no
    for key in result:
        if not result[key]:
            result[key] = "no"

    return result


# ============================================================
# Markdown 输出
# ============================================================

def build_md(
    source_file: str,
    actual_approval: dict,
    step61: dict,
    candidates: dict,
    affected_docs: dict,
) -> str:
    lines = [
        STRUCTURED_SECTION_HEADING,
        "",
        "## structured_fields_actual_approval",
        "",
    ]

    for key in APPROVAL_RESULT_KEYS:
        value = actual_approval.get(key, "")
        if not is_valid_person_name(value):
            value = ""
        lines.append(f"{key}: {value}")

    if has_any_valid_person(actual_approval):
        note = "已从实际 Step 4 / Step 7 签字区域读取到最终签字人。"
    else:
        note = "未在实际 Step 4 / Step 7 签字区域读取到最终签字人；Signature Matrix 仅作为候选人矩阵，不直接填入 approval_person。"

    lines += [
        "",
        f"approval_source_note: {note}",
        "",
        "## structured_fields_step_6_1_responsible",
        "",
    ]

    for key in [
        "implementation_development_responsible",
        "implementation_purchasing_responsible",
        "implementation_mfe_responsible",
        "implementation_cos_responsible",
        "implementation_quality_responsible",
        "implementation_cpjm_responsible",
        "implementation_moex_responsible",
        "implementation_log_responsible",
    ]:
        value = step61.get(key, "")
        if not is_valid_person_name(value):
            value = ""
        lines.append(f"{key}: {value}")

    lines += [
        "",
        "## structured_fields_signature_matrix_candidates",
        "",
    ]

    for key in [
        "development_candidates",
        "purchasing_candidates",
        "mfe_candidates",
        "cos_candidates",
        "quality_candidates",
        "cpjm_candidates",
        "moex_candidates",
        "log_candidates",
    ]:
        values = candidates.get(key, [])
        values = [v for v in values if is_valid_person_name(v)]
        values = unique_keep_order(values)
        lines.append(f"{key}: {'; '.join(values)}")

    lines += [
        "",
        "## structured_fields_step_3_3_affected_documents",
        "",
    ]

    for key in [
        "interface_fmea_value",
        "product_fmea_value",
        "special_characteristics_value",
        "imds_value",
        "offer_drawing_value",
        "tcd_value",
        "norm_wb_hf_value",
        "affected_document_other_value",
    ]:
        value = affected_docs.get(key, "no")
        if value not in ["yes", "no"]:
            value = "no"
        lines.append(f"{key}: {value}")

    return "\n".join(lines) + "\n"


def remove_existing_structured_section(text: str) -> str:
    marker = f"\n{STRUCTURED_SECTION_HEADING}"
    marker_index = text.find(marker)

    if marker_index < 0:
        if text.startswith(STRUCTURED_SECTION_HEADING):
            return ""
        return text.rstrip()

    return text[:marker_index].rstrip()


def merge_structured_fields_into_case_md(file_path: Path, structured_md: str):
    target_path = OUT_DIR / f"{file_path.stem}.md"

    if target_path.exists():
        existing_text = target_path.read_text(encoding="utf-8", errors="ignore")
    else:
        existing_text = "\n".join(
            [
                "# Historical PD-ECR Case",
                f"Source file: {file_path.name}",
                "",
            ]
        )

    base_text = remove_existing_structured_section(existing_text)
    merged_text = f"{base_text}\n\n{structured_md}".strip() + "\n"
    target_path.write_text(merged_text, encoding="utf-8")

    return target_path


# ============================================================
# 主流程
# ============================================================

def process_one_excel(file_path: Path):
    print("=" * 80)
    print(f"处理: {file_path.name}")

    wb = load_workbook(file_path, data_only=True)

    actual_approval = extract_actual_approval_from_workbook(wb)
    step61 = extract_step61_responsible_from_workbook(wb)
    candidates = extract_signature_matrix_candidates_from_workbook(wb)
    affected_docs = extract_step33_affected_documents_from_workbook(wb)

    md = build_md(
        source_file=file_path.name,
        actual_approval=actual_approval,
        step61=step61,
        candidates=candidates,
        affected_docs=affected_docs,
    )

    out_path = merge_structured_fields_into_case_md(file_path, md)

    print("合并结构化签字字段:", out_path)
    print("actual_approval:", actual_approval)
    print("step6.1:", step61)
    print("candidates:", candidates)
    print("affected_docs:", affected_docs)


def main():
    excel_files = (
        list(EXCEL_DIR.glob("*.xlsx"))
        + list(EXCEL_DIR.glob("*.xlsm"))
    )

    print(f"共发现 {len(excel_files)} 个 Excel 文件")
    print("输入目录:", EXCEL_DIR)
    print("输出目录:", OUT_DIR)

    for file_path in excel_files:
        try:
            process_one_excel(file_path)
        except Exception as e:
            print(f"[ERROR] 处理失败: {file_path.name}")
            print(e)

    print("全部完成")


if __name__ == "__main__":
    main()
