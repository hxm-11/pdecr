from pathlib import Path
import sys
from openpyxl import load_workbook

try:
    import xlrd
except ImportError:
    xlrd = None

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


BASE_DIR = Path(__file__).resolve().parent
EXCEL_DIR = BASE_DIR / "excel_source"
KNOWLEDGE_DIR = BASE_DIR / "knowledge"

KNOWLEDGE_DIR.mkdir(parents=True, exist_ok=True)


KEYWORDS = [
    "DC No",
    "Date",
    "Customer project",
    "Customer project Name",
    "MCR No",
    "Product No",
    "Component No",
    "Initiator",
    "Reason of changes",
    "Reason of change",
    "Current design",
    "Change proposal",
    "Remarks",
    "Impact analysis",
    "Function & Performance",
    "Interface and Appearance",
    "Reliability and robustness",
    "Manufacturing",
    "assembly",
    "testing",
    "supplier part",
    "System",
    "Hardware",
    "Software",
    "Calibration",
    "Quality Assurance",
    "Trial run",
    "Capability",
    "CMK",
    "MSA",
    "MAE release",
    "BOM check",
    "Test report",
    "Implementation",
    "Approval",
    "Development",
    "Purchasing",
    "MFE",
    "Quality",
    "COS",
    "MOEx",
    "LOG",
]


def clean(value):
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    text = " ".join(text.split())
    return text


def row_to_text(row_values):
    values = [clean(v) for v in row_values]
    values = [v for v in values if v]

    if not values:
        return ""

    return " | ".join(values)


def is_relevant_row(text: str) -> bool:
    lower_text = text.lower()

    for kw in KEYWORDS:
        if kw.lower() in lower_text:
            return True

    # 保留有勾选框或 Y/N 的行
    if "☑" in text or "☐" in text:
        return True

    if " yes " in f" {lower_text} " or " no " in f" {lower_text} ":
        return True

    return False


def build_output_from_sheets(source_file: str, sheets: list[tuple[str, list[list[object]]]]):
    output = []
    output.append("# Historical PD-ECR Case")
    output.append(f"Source file: {source_file}")
    output.append("")
    output.append("## File Type")
    output.append("Excel")
    output.append("")

    for sheet_name, rows in sheets:
        sheet_lines = []

        for row in rows:
            line = row_to_text(row)

            if not line:
                continue

            if is_relevant_row(line):
                sheet_lines.append(line)

        if sheet_lines:
            output.append(f"## Sheet: {sheet_name}")
            output.extend(sheet_lines)
            output.append("")

    return output


def read_win32com_sheets(file_path: Path):
    """Read Excel via COM automation, capturing Form Control checkbox states
    that openpyxl cannot see.  Returns the same (sheet_name, rows) shape as
    read_openpyxl_sheets, with checkbox captions/states appended as extra rows.
    """
    import pythoncom

    pythoncom.CoInitialize()
    try:
        try:
            import win32com.client
        except ImportError:
            return None  # caller should fall back to openpyxl

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(file_path.absolute()))
            sheets = []
            for ws in wb.Worksheets:
                rows: list[list[object]] = []

                # ── Read cell values ──
                used = ws.UsedRange
                if used is not None:
                    raw = used.Value
                    if raw is None:
                        pass  # empty sheet
                    elif isinstance(raw, tuple):
                        for row_tuple in raw:
                            if isinstance(row_tuple, tuple):
                                rows.append(list(row_tuple))
                            else:
                                rows.append([row_tuple])
                    else:
                        rows.append([raw])

                # ── Read Form Control checkboxes ──
                checkbox_rows = _read_checkboxes(ws)
                if checkbox_rows:
                    # Insert a marker row before appending checkbox data
                    if rows:
                        rows.append([""] * max(len(r) for r in rows))
                    for cb_row in checkbox_rows:
                        rows.append(cb_row)

                sheets.append((ws.Name, rows))

            wb.Close(False)
            return sheets
        finally:
            excel.Quit()
    finally:
        pythoncom.CoUninitialize()


def _read_checkboxes(ws) -> list[list[object]]:
    """Extract Form Control checkbox captions and states from a worksheet."""
    rows: list[list[object]] = []
    try:
        shapes = ws.CheckBoxes()
    except Exception:
        # Fallback: try OLEObjects or Shapes
        try:
            shapes = ws.OLEObjects()
        except Exception:
            try:
                shapes = ws.Shapes()
            except Exception:
                return rows

    count = shapes.Count if hasattr(shapes, "Count") else 0
    for i in range(1, count + 1):
        try:
            cb = shapes.Item(i)
        except Exception:
            continue

        caption = ""
        checked = False
        cell_ref = ""

        try:
            caption = str(cb.Caption or cb.Text or cb.Name or "").strip()
        except Exception:
            try:
                caption = str(cb.Name or "").strip()
            except Exception:
                pass

        try:
            # 1 = checked, -4146 (xlOff) or 0 = unchecked
            val = cb.Value
            checked = val == 1
        except Exception:
            pass

        try:
            cell_ref = str(cb.LinkedCell or cb.TopLeftCell.Address or "")
        except Exception:
            pass

        if caption:
            # Place in the approximate column position based on the control's location
            # Use a fixed-width row: [label, status, cell_ref]
            marker = "☑" if checked else "☐"
            row = [caption, marker, cell_ref] if cell_ref else [caption, marker]
            rows.append(row)

    return rows


def read_openpyxl_sheets(file_path: Path):
    wb = load_workbook(file_path, data_only=True)
    return [
        (sheet_name, list(wb[sheet_name].iter_rows(values_only=True)))
        for sheet_name in wb.sheetnames
    ]


def read_xlrd_sheets(file_path: Path):
    if xlrd is None:
        raise RuntimeError("读取 .xls 需要安装 xlrd：python -m pip install xlrd")

    book = xlrd.open_workbook(str(file_path))
    sheets = []

    for sheet in book.sheets():
        rows = [
            sheet.row_values(row_index)
            for row_index in range(sheet.nrows)
        ]
        sheets.append((sheet.name, rows))

    return sheets


def convert_excel(file_path: Path):
    suffix = file_path.suffix.lower()

    if suffix == ".xls":
        sheets = read_xlrd_sheets(file_path)
    else:
        # Prefer win32com — it captures checkbox states that openpyxl misses
        sheets = read_win32com_sheets(file_path)
        if sheets is None:
            sheets = read_openpyxl_sheets(file_path)

    output = build_output_from_sheets(file_path.name, sheets)

    out_path = KNOWLEDGE_DIR / f"{file_path.stem}.md"
    out_path.write_text("\n".join(output), encoding="utf-8")

    print(f"已转换：{file_path.name} -> {out_path.name}")


def main():
    files = (
        list(EXCEL_DIR.glob("*.xlsx"))
        + list(EXCEL_DIR.glob("*.xlsm"))
        + list(EXCEL_DIR.glob("*.xls"))
    )

    if not files:
        print(f"没有找到 Excel 文件，请放到：{EXCEL_DIR}")
        return

    for f in files:
        if f.name.startswith("~$"):
            continue

        try:
            convert_excel(f)
        except Exception as e:
            print(f"转换失败：{f.name}，错误：{e}")


if __name__ == "__main__":
    main()
