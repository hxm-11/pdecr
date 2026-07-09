"""Excel (.xlsx / .xls) -> ParsedDocument。

每个 sheet 转成一个 ParsedTable（二维单元格 + 渲染文本），并把所有 sheet
的文本拼成正文。Excel 内容不会直接进向量库——它和其它来源一样，必须先经
抽取器进入统一的 PdecrCase schema。

优先用 openpyxl（.xlsx）；.xls 走 pandas（需 xlrd）。都不可用时给出明确报错。
"""

from __future__ import annotations

from pathlib import Path

from ..loaders import ParsedDocument, ParsedTable, compute_checksum


class ExcelParser:
    parser_name = "excel"

    def parse(self, excel_path: str) -> ParsedDocument:
        path = Path(excel_path)
        if not path.exists():
            raise FileNotFoundError(excel_path)

        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            tables = self._parse_with_openpyxl(path)
        else:  # .xls 等
            tables = self._parse_with_pandas(path)

        text = "\n\n".join(t.text for t in tables if t.text)

        return ParsedDocument(
            source_file=path.name,
            file_type=suffix.lstrip("."),
            parser=self.parser_name,
            text=text,
            tables=tables,
            checksum=compute_checksum(path),
        )

    # ── openpyxl ──────────────────────────────────────────
    @staticmethod
    def _parse_with_openpyxl(path: Path) -> list[ParsedTable]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError(
                "需要 openpyxl 读取 .xlsx：pip install openpyxl"
            ) from exc

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        tables: list[ParsedTable] = []
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in row]
                if any(cells):
                    rows.append(cells)
            tables.append(_rows_to_table(ws.title, rows))
        wb.close()
        return tables

    # ── pandas（.xls 回退）─────────────────────────────────
    @staticmethod
    def _parse_with_pandas(path: Path) -> list[ParsedTable]:
        try:
            import pandas as pd
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError(
                "需要 pandas 读取 .xls：pip install pandas xlrd"
            ) from exc

        sheets = pd.read_excel(str(path), sheet_name=None, header=None, dtype=str)
        tables: list[ParsedTable] = []
        for name, df in sheets.items():
            df = df.fillna("")
            rows = [[str(c).strip() for c in row] for row in df.values.tolist()]
            rows = [r for r in rows if any(r)]
            tables.append(_rows_to_table(str(name), rows))
        return tables


def _rows_to_table(name: str, rows: list[list[str]]) -> ParsedTable:
    lines = [f"# Sheet: {name}"]
    for row in rows:
        # 用制表符渲染，保留空单元占位以维持列对齐
        lines.append("\t".join(row))
    return ParsedTable(name=name, rows=rows, text="\n".join(lines))
